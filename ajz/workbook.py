"""Workbook generator (spec §9).

Writes the finished Excel file Jeff opens. Design rules, all of them reactions to how
v5.1 failed in his hands:

* **Every cell is a static value.** No formulas, no cross-sheet references. Nothing in
  the file can go #REF!, and nothing recalculates behind him. The generator is the only
  thing that computes.
* **Only two sheets are editable** — Universe and Settings. The rest are protected, so
  he cannot accidentally type over his own dashboard.
* **Data validation on every score cell**, so a typo is refused at entry rather than
  silently corrupting a conviction score.
* **Nothing is written for a row that does not exist.** v5.1 pre-filled 499 empty rows
  with formulas that returned 0; every average then divided by 499 and read ~0 forever.
"""

from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from . import __version__, palette, theme
from .bands import BandTable
from .calc import rank_pre_profit, rank_stocks
from .models import ScoredStock
from .settings import (
    BAND_TABLES,
    DEFAULT_THRESHOLDS,
    SPARE_BAND_ROWS,
    TABLE_END,
    TABLE_PREFIX,
    Thresholds,
)
from .status import RefreshStatus

THIN = Side(style="thin", color=theme.RULE)
CELL_BORDER = Border(bottom=THIN)


def _fill(argb: str | None) -> PatternFill | None:
    if argb is None:
        return None
    return PatternFill("solid", fgColor=argb)


def _write_header(ws: Worksheet, row: int, headers: list[str], widths: list[int]) -> None:
    for col, (label, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(name=theme.FONT, bold=True, color=theme.INK_INVERSE, size=11)
        cell.fill = _fill(theme.HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 28


def _style_body(cell, *, bold: bool = False, ink: str = theme.INK_PRIMARY,
                number_format: str | None = None, align: str = "left") -> None:
    cell.font = Font(name=theme.FONT, bold=bold, color=ink, size=11)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = CELL_BORDER
    if number_format:
        cell.number_format = number_format


def _protect(ws: Worksheet, allow_formatting: bool = False) -> None:
    """Read-only, but with no password.

    A password would produce a prompt, and a prompt is a thing Jeff has to understand.
    This only stops accidental typing, which is the actual risk.

    `allow_formatting` is for the Settings sheet, and it is not a nicety. In OOXML a
    sheetProtection attribute of "1" means the feature is PROTECTED, not permitted, and
    formatCells defaults to "1". So protecting a sheet at all silently refuses every
    fill command on it -- including on the unlocked cells we specifically asked Jeff to
    colour. He reported that twice as "the colour didn't stick", and he was right both
    times: Excel would not let him make the edit the feature is built on.

    Permitting formatting cannot cost us anything. Formatting changes no value, the
    sheet is rebuilt every refresh anyway, and the fill is now an input we read back.
    """
    # Note: never assign `protection.password`, not even None — openpyxl runs the
    # legacy hasher on whatever it is given and blows up on None. Leaving it unset is
    # what produces the passwordless protection we actually want.
    ws.protection.sheet = True
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    if allow_formatting:
        ws.protection.formatCells = False


# --- Sheet 1: Dashboard ---------------------------------------------------------------


def _build_dashboard(ws: Worksheet, stocks: list[ScoredStock], status: RefreshStatus) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 44

    title = ws.cell(row=2, column=2, value="AJZ Dashboard")
    title.font = Font(name=theme.FONT, bold=True, size=24, color=theme.INK_PRIMARY)

    # --- Status banner. The whole error-reporting surface (spec §10).
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)
    banner = ws.cell(row=4, column=2, value=status.headline)
    banner.fill = _fill(theme.BANNER_FILL[status.state.value])
    banner.font = Font(
        name=theme.FONT, bold=True, size=12, color=theme.BANNER_INK[status.state.value]
    )
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 30

    row = 5
    if status.note:
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)
        note = ws.cell(row=5, column=2, value=status.note)
        note.font = Font(name=theme.FONT, size=10, color=theme.INK_SECONDARY, italic=True)
        note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row = 6

    # --- Deliberately empty below this line.
    #
    # Jeff, v2.1: "I don't think the dashboard does anything at this point. Maybe someday
    # it could be a pie chart summary of portfolio. I would eliminate the data but leave
    # the sheet for future use."
    #
    # So the tiles are gone and the sheet stays. What remains is the status banner above,
    # which is not "data" in the sense he meant -- it is the entire error-reporting
    # surface (spec §10), the one place that says whether the numbers he is about to read
    # arrived today or are three days stale. Deleting it would leave a silent failure
    # with nowhere to appear, which is how v5.1 managed to show zeros for months.
    row += 2
    placeholder = ws.cell(
        row=row, column=2,
        value="This sheet is reserved for a future portfolio summary.",
    )
    placeholder.font = Font(name=theme.FONT, size=11, italic=True, color=theme.INK_MUTED)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    row += 2
    hint = ws.cell(
        row=row, column=2,
        value="Your rankings are on the Top Rankings sheet. To add or remove a stock, "
              "use the Universe sheet. To change how the categories are set, use "
              "Settings.",
    )
    hint.font = Font(name=theme.FONT, size=10, italic=True, color=theme.INK_MUTED)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    # In the file rather than only in a log, because the file is the thing Jeff has in
    # front of him when he emails to say something looks wrong.
    row += 2
    stamp = ws.cell(row=row, column=2, value=f"AJZ Dashboard v{__version__}")
    stamp.font = Font(name=theme.FONT, size=9, color=theme.INK_MUTED)

    _protect(ws)


# --- Sheet 2: Top Rankings ------------------------------------------------------------


def _build_rankings(ws: Worksheet, stocks: list[ScoredStock],
                    thresholds: Thresholds) -> None:
    # Column order is Jeff's, from v2.1: each number is immediately followed by the word
    # for it, so he reads across a row rather than back and forth to a legend. His
    # Conviction and Category columns (old H, I, J) are gone with the feature.
    headers = ["Rank", "Ticker", "Company", "Sector",
               "AJZ Score", "Score Category",
               "Forward P/E", "P/E Category",
               "AJZ Value", "Value Category",
               "Rank Δ", "Notes"]
    widths = [7, 10, 28, 20, 11, 15, 12, 17, 11, 15, 8, 40]
    _write_header(ws, 1, headers, widths)

    ranked = rank_stocks(stocks)
    pre_profit = rank_pre_profit(stocks)
    unscored = [s for s in stocks if not s.is_rankable and not s.is_pre_profit]

    row = 2
    for position, s in enumerate(ranked, start=1):
        _write_ranking_row(ws, row, s, position, thresholds)
        row += 1

    # Pre-profit companies, ranked on AJZ Score. Jeff asked for these to be included
    # rather than merely listed -- he expects to invest in some of them -- but was clear
    # that they "shouldn't pollute the others", so they sit under their own heading with
    # their own numbering and stay out of every average.
    #
    # The rank reads "P1", not "1". Two orderings on one sheet in one column is exactly
    # the sort of thing that gets read as one ordering, and these are not comparable:
    # the number above is an AJZ Value Score, the number here is an AJZ Score.
    if pre_profit:
        row += 1
        heading = ws.cell(
            row=row, column=1,
            value=f"{thresholds.pre_profit_label} — ranked on AJZ Score only "
                  f"(no forward P/E, and never included in any average)")
        heading.font = Font(name=theme.FONT, bold=True, size=11,
                            color=theme.INK_SECONDARY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        row += 1
        for position, s in enumerate(pre_profit, start=1):
            _write_ranking_row(ws, row, s, None, thresholds, rank_text=f"P{position}")
            row += 1

    # Everything else: rows with no AJZ Score at all. Missing data, not missing
    # earnings, so there is nothing honest to rank them on.
    if unscored:
        row += 1
        heading = ws.cell(row=row, column=1,
                          value="Not scored — see the Notes column for what is missing")
        heading.font = Font(name=theme.FONT, bold=True, size=11, color=theme.INK_MUTED)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        row += 1
        for s in unscored:
            _write_ranking_row(ws, row, s, None, thresholds)
            row += 1

    ws.freeze_panes = "C2"
    # The filter covers the main ranking only. Extending it over the section headings
    # below would let a filter interleave three lists that are ordered on three
    # different quantities — which is the exact confusion the headings exist to prevent.
    last_ranked = 1 + len(ranked)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(last_ranked, 1)}"
    _protect(ws)


def _band_paint(table: BandTable | None, label: str | None) -> tuple[str | None, str]:
    """(fill, ink) for a band: Jeff's own colour if he set one, otherwise our ramp.

    The single place colour is decided, so the Settings legend, the Top Rankings columns
    and the Opportunity Matrix cannot drift apart. That mattered as soon as he started
    picking colours: three surfaces each computing their own shading is three chances for
    the same category to appear in three different colours.

    Ink follows the fill rather than being chosen alongside it, because he picks the fill
    and nobody should have to also pick readable text to go on it.
    """
    index = table.shade_index(label) if table is not None else None
    if index is None:
        return None, theme.INK_MUTED

    band = table.band_for(label)
    if band is not None and band.color:
        return band.color, palette.ink_for(band.color)
    return theme.band_style(index, len(table.bands))


def _banded(ws: Worksheet, row: int, col: int, label: str | None,
            table: BandTable | None = None) -> None:
    """A category cell: the word, shaded by its band, never colour alone.

    Passing the table is what lets a renamed band keep its shading; without it the cell
    still says the right word, just without the colour.
    """
    cell = ws.cell(row=row, column=col, value=label or "—")
    fill, ink = _band_paint(table, label)
    _style_body(cell, bold=True, align="center", ink=ink)
    if fill:
        cell.fill = _fill(fill)


def _write_ranking_row(ws: Worksheet, row: int, s: ScoredStock,
                       position: int | None, thresholds: Thresholds,
                       rank_text: str | None = None) -> None:
    d = s.data
    rank_cell = ws.cell(row=row, column=1,
                        value=rank_text or (position if position else "—"))
    _style_body(rank_cell, bold=True, align="center",
                ink=theme.INK_PRIMARY if (position or rank_text) else theme.INK_MUTED)

    _style_body(ws.cell(row=row, column=2, value=d.ticker), bold=True)
    _style_body(ws.cell(row=row, column=3, value=d.company or ""))
    _style_body(ws.cell(row=row, column=4, value=d.sector or ""), ink=theme.INK_SECONDARY)

    _style_body(ws.cell(row=row, column=5,
                        value=s.ajz_score if s.ajz_score is not None else "—"),
                number_format="0.0" if s.ajz_score is not None else None, align="right")
    _banded(ws, row, 6, s.score_label, thresholds.score_bands)

    _style_body(ws.cell(row=row, column=7,
                        value=s.forward_pe if s.forward_pe is not None else "—"),
                number_format="0.0" if s.forward_pe is not None else None, align="right")
    _banded(ws, row, 8, s.pe_label, thresholds.pe_bands)

    _style_body(ws.cell(row=row, column=9,
                        value=s.ajz_value_score if s.ajz_value_score is not None else "—"),
                bold=True,
                number_format="0.00" if s.ajz_value_score is not None else None,
                align="right")
    _banded(ws, row, 10, s.value_label, thresholds.value_bands)

    _style_body(ws.cell(row=row, column=11, value="—"), align="center", ink=theme.INK_MUTED)
    _style_body(ws.cell(row=row, column=12, value="; ".join(s.notes)), ink=theme.INK_MUTED)


# --- Sheet 3: Opportunity Matrix ------------------------------------------------------


def _build_matrix(ws: Worksheet, stocks: list[ScoredStock],
                  thresholds: Thresholds) -> None:
    """The Primary Screen, as columns of tickers under each of Jeff's Value categories.

    Was a 2x2 of AJZ Value against Conviction. Jeff's v2.1: "I think Opportunity Matrix
    Sheet should be the categories of Primary Screen". With conviction gone one axis of
    the 2x2 no longer exists, and a 2x2 with one real axis is a list drawn as a square.

    So it is a list — one column per band, in his order, best on the left. He also wrote
    "For future there may be a Secondary Screen or portfolio calc but for now I think I
    need to keep it simple", which is an instruction not to invent a second axis to
    replace the one we removed.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    title = ws.cell(row=2, column=2, value="Opportunity Matrix")
    title.font = Font(name=theme.FONT, bold=True, size=20)

    sub = ws.cell(row=3, column=2, value="Primary Screen — AJZ Value Score categories")
    sub.font = Font(name=theme.FONT, size=11, italic=True, color=theme.INK_SECONDARY)

    ranked = rank_stocks(stocks)
    pre_profit = rank_pre_profit(stocks)
    bands = thresholds.value_bands.bands
    ranges = thresholds.value_bands.display_ranges()

    for index, (band, range_text) in enumerate(zip(bands, ranges)):
        col = 2 + index
        ws.column_dimensions[get_column_letter(col)].width = 20

        head = ws.cell(row=5, column=col, value=band.label)
        fill, ink = _band_paint(thresholds.value_bands, band.label)
        head.fill = _fill(fill or theme.NEUTRAL_FILL)
        head.font = Font(name=theme.FONT, bold=True, size=12, color=ink)
        head.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 24

        cap = ws.cell(row=6, column=col, value=range_text)
        cap.font = Font(name=theme.FONT, size=9, italic=True, color=theme.INK_MUTED)
        cap.alignment = Alignment(horizontal="center")

        members = [s for s in ranked if s.value_label == band.label]
        for offset, stock in enumerate(members):
            cell = ws.cell(row=7 + offset, column=col,
                           value=f"{stock.data.ticker}   {stock.ajz_value_score:.1f}")
            cell.font = Font(name=theme.FONT, size=11)
            cell.alignment = Alignment(horizontal="center")

        # An empty category is information, not a gap: it says his cut-off is above
        # everything he owns. Left unlabelled it reads as a rendering failure instead.
        if not members:
            empty = ws.cell(row=7, column=col, value="none")
            empty.font = Font(name=theme.FONT, size=10, italic=True,
                              color=theme.INK_MUTED)
            empty.alignment = Alignment(horizontal="center")

    # The pre-profit column. Deliberately last and deliberately neutral-coloured: it is
    # not a rung on the Value ladder, it is the set of companies the ladder cannot
    # measure. Giving it a band colour would place it in an order it has no place in.
    if pre_profit:
        col = 2 + len(bands)
        ws.column_dimensions[get_column_letter(col)].width = 20

        head = ws.cell(row=5, column=col, value=thresholds.pre_profit_label)
        head.fill = _fill(theme.NEUTRAL_FILL)
        head.font = Font(name=theme.FONT, bold=True, size=12, color=theme.INK_SECONDARY)
        head.alignment = Alignment(horizontal="center", vertical="center")

        cap = ws.cell(row=6, column=col, value="AJZ Score — no P/E")
        cap.font = Font(name=theme.FONT, size=9, italic=True, color=theme.INK_MUTED)
        cap.alignment = Alignment(horizontal="center")

        # The AJZ Score, not the Value Score, and labelled as such in the caption above.
        # A bare number under a column of Value Scores would be read as one.
        for offset, stock in enumerate(pre_profit):
            cell = ws.cell(row=7 + offset, column=col,
                           value=f"{stock.data.ticker}   {stock.ajz_score:.0f}")
            cell.font = Font(name=theme.FONT, size=11)
            cell.alignment = Alignment(horizontal="center")

    unscored = [s for s in stocks if not s.is_rankable and not s.is_pre_profit]
    if unscored:
        row = 9 + max([sum(1 for s in ranked if s.value_label == b.label)
                       for b in bands] + [len(pre_profit)], default=0)
        note = ws.cell(row=row, column=2,
                       value=f"Not scored: {', '.join(s.ticker for s in unscored)} — "
                             "see the Notes column on Top Rankings for what is missing.")
        note.font = Font(name=theme.FONT, size=10, italic=True, color=theme.INK_MUTED)

    _protect(ws)


def _build_alerts(ws: Worksheet, stocks: list[ScoredStock],
                  thresholds: Thresholds) -> None:
    headers = ["Alert", "Ticker", "Company", "AJZ Value", "Category", "Why"]
    widths = [14, 10, 30, 12, 15, 58]
    _write_header(ws, 1, headers, widths)

    severity = {"EXIT": 0, "WARNING": 1, "DOWNGRADE": 2, "BUY": 3, "UPGRADE": 4}
    rows: list[tuple[str, ScoredStock]] = [
        (a.value, s) for s in stocks for a in s.alerts
    ]
    rows.sort(key=lambda r: (severity.get(r[0], 9), -(r[1].ajz_value_score or 0)))

    if not rows:
        cell = ws.cell(row=2, column=1, value="No alerts today.")
        cell.font = Font(name=theme.FONT, italic=True, color=theme.INK_MUTED)
        _protect(ws)
        return

    for row, (alert, s) in enumerate(rows, start=2):
        badge = ws.cell(row=row, column=1, value=alert)
        _style_body(badge, bold=True, align="center", ink=theme.ALERT_INK[alert])
        badge.fill = _fill(theme.ALERT_FILL[alert])

        _style_body(ws.cell(row=row, column=2, value=s.data.ticker), bold=True)
        _style_body(ws.cell(row=row, column=3, value=s.data.company or ""))
        _style_body(ws.cell(row=row, column=4, value=s.ajz_value_score),
                    number_format="0.00", align="right")
        _banded(ws, row, 5, s.value_label, thresholds.value_bands)
        _style_body(ws.cell(row=row, column=6, value=_explain(alert, s, thresholds)),
                    ink=theme.INK_SECONDARY)

    ws.freeze_panes = "A2"
    _protect(ws)


def _explain(alert: str, s: ScoredStock, thresholds: Thresholds) -> str:
    """Every alert says why, in words, quoting the number he set.

    The thresholds are read rather than hardcoded so the sentence cannot drift from the
    rule. The previous version wrote "is above 7" as literal text, which would have
    quietly started lying the first time Jeff changed 7 to something else on the
    Settings sheet -- a wrong explanation is worse than none, because he would trust it.
    """
    v = s.ajz_value_score
    if alert == "BUY":
        return f"AJZ Value {v:.1f} is above {thresholds.buy_value:g}."
    if alert == "WARNING":
        return f"AJZ Value {v:.1f} has fallen below {thresholds.warning_value:g}."
    if alert == "EXIT":
        return f"AJZ Value {v:.1f} is below {thresholds.exit_value:g}."
    if alert == "UPGRADE":
        return (f"AJZ Score up more than {thresholds.mover_score_pct:g}%, forward P/E "
                f"down more than {thresholds.mover_pe_pct:g}%, or moved up a category.")
    if alert == "DOWNGRADE":
        return (f"AJZ Score down more than {thresholds.mover_score_pct:g}% or forward "
                f"P/E up more than {thresholds.mover_pe_pct:g}%.")
    return ""


# --- Sheet 5: Movers ------------------------------------------------------------------


def _build_movers(ws: Worksheet, stocks: list[ScoredStock],
                  thresholds: Thresholds, movement: dict | None = None) -> None:
    """What has actually changed since the last refresh.

    Jeff, v2.1: "Please check the Movers page as it doesn't look like its updating."
    He was right, and it was worse than not updating: the sheet was a stub. It wrote one
    italic line saying rank changes would appear after the second refresh and then never
    wrote anything again, no matter how many refreshes ran. `History.movers()` existed
    and was tested; nothing ever called it into the workbook.

    He also specified the rules and the empty state: alert on an AJZ Score move over 25%
    or a forward P/E move over 10%, and "maybe there should be a note that says no movers
    if no stock has moved to a different category".

    Three distinct empty states, kept distinct. "No baseline yet" is not "nothing moved",
    and neither is "nothing crossed a category line" -- collapsing them is how a sheet
    ends up quietly reporting calm while it is in fact blind.
    """
    headers = ["Ticker", "Company", "What moved", "Was", "Now", "Change"]
    widths = [10, 28, 20, 14, 14, 12]
    _write_header(ws, 1, headers, widths)

    movement = movement or {}
    rows = movement.get("rows") or []

    if not movement.get("has_baseline"):
        _note(ws, 2, len(headers),
              "This is the first refresh, so there is nothing to compare against yet. "
              "Movers will start reporting from your next refresh.")
        _protect(ws)
        return

    if not rows:
        _note(ws, 2, len(headers),
              f"No movers. Nothing has moved more than "
              f"{thresholds.mover_score_pct:g}% on AJZ Score or "
              f"{thresholds.mover_pe_pct:g}% on forward P/E, and no stock has changed "
              "category on AJZ Score, Forward P/E or AJZ Value Score.")
        _protect(ws)
        return

    for row, entry in enumerate(rows, start=2):
        _style_body(ws.cell(row=row, column=1, value=entry["ticker"]), bold=True)
        _style_body(ws.cell(row=row, column=2, value=entry.get("company") or ""))
        _style_body(ws.cell(row=row, column=3, value=entry["what"]),
                    ink=theme.INK_SECONDARY)
        _style_body(ws.cell(row=row, column=4, value=entry["was"]), align="right")
        _style_body(ws.cell(row=row, column=5, value=entry["now"]), align="right")

        change = ws.cell(row=row, column=6, value=entry["change"])
        _style_body(change, bold=True, align="right",
                    ink=theme.INK_POSITIVE if entry["improved"] else theme.INK_NEGATIVE)

    ws.freeze_panes = "A2"
    _protect(ws)


def _note(ws: Worksheet, row: int, width: int, text: str) -> None:
    """A merged italic line explaining why a sheet is empty.

    An unexplained blank sheet reads as broken -- Jeff shipped three of them in v5.1 and
    could not tell which were failing and which simply had nothing to say.
    """
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=theme.FONT, italic=True, color=theme.INK_MUTED)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    ws.row_dimensions[row].height = 32


# --- Sheet 8: Settings (EDITABLE) -----------------------------------------------------


def _build_settings(ws: Worksheet, stocks: list[ScoredStock],
                    thresholds: Thresholds) -> None:
    """Everything Jeff can change without calling anyone.

    Deliberately NOT here: the AJZ Score weights. Those are AJZ Rule 3.0 itself — he
    wrote "Keep Unchanged" beside them, and altering them changes what a score means,
    breaking comparability with every stored snapshot. What IS here is everything that
    turns a number into a word, which is investment judgement and therefore his.

    Column D holds the field key and is hidden: the read-back matches on it rather than
    on the label text, so re-wording a label can never silently orphan a setting — and he
    is expected to re-word labels, having renamed two of his own bands between v2.0 and
    v2.1 of the change request.
    """
    widths = [46, 13, 34, 3, 13]
    _write_header(ws, 1, ["Setting", "Your value", "What it does"], widths)
    ws.column_dimensions["D"].hidden = True

    intro = ws.cell(row=1, column=6,
                    value="Change anything in the shaded columns and it takes effect at "
                          "the next refresh. Clear a cell to restore its default. "
                          "Fill a category name cell with a colour and that category "
                          "takes the colour everywhere it appears.")
    intro.font = Font(name=theme.FONT, size=10, italic=True, color=theme.INK_SECONDARY)

    positive = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1=0, allow_blank=True,
        showErrorMessage=True, errorTitle="Must be zero or more",
        error="Enter a number, or clear the cell to use the default.",
    )
    ws.add_data_validation(positive)

    row = 2
    for key, label, value, explanation in thresholds.describe():
        _style_body(ws.cell(row=row, column=1, value=label))
        cell = ws.cell(row=row, column=2, value=value)
        _style_body(cell, bold=True, align="center")
        cell.protection = Protection(locked=False)
        # Only the numeric settings get the numeric validator. Attaching it to a text
        # setting would make Excel reject the word he is supposed to type there.
        if not isinstance(value, str):
            positive.add(cell)
        _style_body(ws.cell(row=row, column=3, value=explanation), ink=theme.INK_MUTED)
        ws.cell(row=row, column=4, value=key)
        row += 1

    for attr, title, _default in BAND_TABLES:
        row = _write_band_table(ws, row + 1, attr, title,
                                getattr(thresholds, attr), stocks)

    ws.freeze_panes = "A2"
    # The one sheet whose formatting is an input rather than an output.
    _protect(ws, allow_formatting=True)


def _band_counts(attr: str, table: BandTable,
                 stocks: list[ScoredStock]) -> dict[str, int]:
    """How many stocks currently land in each band of a table.

    Shown beside every band because a category nobody falls into is otherwise invisible.
    That is not hypothetical: "Aggressive Position" sat permanently unreachable in the
    previous version and nobody noticed for weeks. It is also the fastest way for Jeff to
    see that his own AJZ Score table puts thirteen of twenty-four stocks in "Legendary" —
    a fact about his numbers that no amount of us explaining lands as well as the count
    sitting next to the word.
    """
    getter = {
        "score_bands": lambda s: s.ajz_score,
        "pe_bands": lambda s: s.forward_pe,
        "value_bands": lambda s: s.ajz_value_score,
    }[attr]

    counts = {b.label: 0 for b in table.bands}
    for stock in stocks:
        label = table.label_for(getter(stock))
        if label is not None:
            counts[label] = counts.get(label, 0) + 1
    return counts


def _write_band_table(ws: Worksheet, row: int, attr: str, title: str,
                      table: BandTable, stocks: list[ScoredStock]) -> int:
    """Write one editable category table. Returns the next free row."""
    heading = ws.cell(row=row, column=1, value=title)
    heading.font = Font(name=theme.FONT, bold=True, size=12)
    ws.cell(row=row, column=4, value=f"{TABLE_PREFIX}{attr}")
    row += 1

    for col, label in enumerate(["Category name — fill the cell to set its colour",
                                 "Starts at", "Range (automatic)",
                                 None, "Stocks now"], start=1):
        if label is None:
            continue
        head = ws.cell(row=row, column=col, value=label)
        head.font = Font(name=theme.FONT, bold=True, size=10, color=theme.INK_SECONDARY)
    row += 1

    floors = DataValidation(
        type="decimal", allow_blank=True, showErrorMessage=True,
        operator="between", formula1=-1000000, formula2=1000000,
        errorTitle="Enter a number",
        error="Type the number this category starts at, or clear both cells to remove it.",
    )
    ws.add_data_validation(floors)

    counts = _band_counts(attr, table, stocks)
    first = row

    # Real bands, then blank spares. Excel refuses to insert a row into a protected
    # sheet, so spares are how he adds a category; the reader sorts by floor, so it does
    # not matter that the only free rows are at the bottom.
    entries = list(table.rows()) + [(None, None, None)] * SPARE_BAND_ROWS
    for label, floor, _colour in entries:
        name = ws.cell(row=row, column=1, value=label)
        # Seeded with the colour this band currently shows, so the table doubles as the
        # legend for the whole workbook: he can see what each category looks like on the
        # sheets that use it, and change it by filling the cell. That is what he tried to
        # do unprompted, and it is the right place for it -- the categories are defined
        # here, so their colours should be too. Spare rows stay unfilled; a band typed
        # into one picks up the ramp until he says otherwise.
        fill, ink = _band_paint(table, label)
        _style_body(name, ink=ink if fill else theme.INK_PRIMARY)
        if fill:
            name.fill = _fill(fill)
        name.protection = Protection(locked=False)

        value = ws.cell(row=row, column=2, value=floor)
        _style_body(value, bold=True, align="center", number_format="General")
        value.protection = Protection(locked=False)
        floors.add(value)

        # A live formula, not text stamped at refresh time. Moving one floor visibly
        # re-shapes the band above it the moment he presses Enter — he sees the
        # consequence of the edit while he is still deciding, with no refresh, no
        # network, and nobody to ask. It reads the floors; the floors never read it.
        if row == first:
            formula = f'=IF(B{row}="","",TEXT(B{row},"General")&" and above")'
        else:
            formula = (
                f'=IF(B{row}="","",'
                f'IF(B{row + 1}="","Below "&TEXT(B{row - 1},"General"),'
                f'TEXT(B{row - 1}-0.1,"General")&" – "&TEXT(B{row},"General")))'
            )
        _style_body(ws.cell(row=row, column=3, value=formula), ink=theme.INK_MUTED)

        count = counts.get(label) if label else None
        cell = ws.cell(row=row, column=5, value=count if count is not None else "")
        _style_body(cell, align="center",
                    ink=theme.INK_MUTED if not count else theme.INK_SECONDARY)
        row += 1

    ws.cell(row=row, column=4, value=TABLE_END)
    return row + 1


# --- Sheet 7: Universe (EDITABLE) -----------------------------------------------------


def _build_universe(ws: Worksheet, stocks: list[ScoredStock]) -> None:
    headers = ["Ticker", "Company", "Sector", "Active", "Notes"]
    widths = [12, 30, 24, 10, 46]
    _write_header(ws, 1, headers, widths)

    intro = ws.cell(row=1, column=7,
                    value="Add a ticker to include it. Set Active to NO to hide one "
                          "without deleting the row.")
    intro.font = Font(name=theme.FONT, size=10, italic=True, color=theme.INK_SECONDARY)

    active_validator = DataValidation(
        type="list", formula1='"YES,NO"', allow_blank=False, showErrorMessage=True,
        errorTitle="Active must be YES or NO", error="Choose YES or NO.",
    )
    ws.add_data_validation(active_validator)

    for row, s in enumerate(sorted(stocks, key=lambda x: x.data.ticker), start=2):
        for col, value in enumerate(
            [s.data.ticker, s.data.company or "", s.data.sector or "", "YES", ""], start=1
        ):
            cell = ws.cell(row=row, column=col, value=value)
            _style_body(cell, bold=(col == 1))
            cell.protection = Protection(locked=False)
        active_validator.add(ws.cell(row=row, column=4))

    # Blank editable rows so adding a stock is obvious rather than a guess.
    for row in range(len(stocks) + 2, len(stocks) + 12):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=False)
            cell.border = CELL_BORDER
        active_validator.add(ws.cell(row=row, column=4))

    ws.freeze_panes = "A2"
    _protect(ws)


# --- Entry point ----------------------------------------------------------------------


def build_workbook(
    stocks: list[ScoredStock],
    status: RefreshStatus | None = None,
    generated_at: datetime | None = None,
    thresholds: Thresholds = DEFAULT_THRESHOLDS,
    movement: dict | None = None,
) -> Workbook:
    """Build the complete workbook. Pure: takes data, returns a Workbook, touches no disk."""
    from .status import RefreshState

    status = status or RefreshStatus(
        state=RefreshState.OK, data_as_of=generated_at or datetime(2026, 8, 19, 6, 5)
    )

    wb = Workbook()
    wb.remove(wb.active)

    builders = [
        ("Dashboard", lambda ws: _build_dashboard(ws, stocks, status)),
        ("Top Rankings", lambda ws: _build_rankings(ws, stocks, thresholds)),
        ("Opportunity Matrix", lambda ws: _build_matrix(ws, stocks, thresholds)),
        ("Alerts", lambda ws: _build_alerts(ws, stocks, thresholds)),
        ("Movers", lambda ws: _build_movers(ws, stocks, thresholds, movement)),
        ("Universe", lambda ws: _build_universe(ws, stocks)),
        ("Settings", lambda ws: _build_settings(ws, stocks, thresholds)),
    ]

    for name, builder in builders:
        builder(wb.create_sheet(title=name))

    return wb
