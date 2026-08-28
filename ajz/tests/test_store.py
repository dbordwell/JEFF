"""Tests for the conviction round-trip — the highest-risk correctness path (spec §7.3).

The governing invariant, restated because everything here serves it:

    NEVER write a workbook containing less conviction data than the one it replaces.

Conviction is the only irreplaceable data in the system. Every test below is either
proving the round-trip preserves it, or proving a failure aborts instead of degrading.
"""

from __future__ import annotations

from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from ajz import theme
from ajz.bands import DEFAULT_SCORE_BANDS
from ajz.fixtures import sample_stocks
from ajz.settings import DEFAULT_THRESHOLDS, from_mapping
from ajz.store import (
    WorkbookLockedError,
    WorkbookReadError,
    archive_conviction,
    atomic_save,
    backup_workbook,
    read_existing,
)
from ajz.workbook import build_workbook


@pytest.fixture
def live_file(tmp_path):
    """A generated workbook on disk, as Jeff would have after one refresh."""
    path = tmp_path / "AJZ Dashboard.xlsx"
    build_workbook(sample_stocks()).save(path)
    return path


def _settings_cell(path, key, value):
    """Simulate Jeff typing into a Settings row identified by its hidden key."""
    wb = load_workbook(path)
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value == key:
            ws.cell(row=row, column=2, value=value)
            break
    else:
        raise AssertionError(f"no Settings row keyed {key!r}")
    wb.save(path)


def _band_rows(path, attr):
    """The (label, floor) rows of one band table, and the sheet rows they occupy."""
    wb = load_workbook(path)
    ws = wb["Settings"]
    start = end = None
    for row in range(2, ws.max_row + 1):
        marker = ws.cell(row=row, column=4).value
        if marker == f"table:{attr}":
            start = row + 2          # skip the title row and the column header
        elif marker == "table:end" and start is not None and end is None:
            end = row
    assert start and end, f"table:{attr} not found on the Settings sheet"
    return ws, start, end


# --- The round-trip that matters now: Jeff's category tables --------------------------


def test_band_tables_written_by_the_generator_read_back_identically(live_file):
    """The narrow contract: what `build_workbook` writes, `read_existing` reads.

    Without this the Settings sheet is decoration -- he edits it, the refresh ignores
    him, and nothing anywhere says so.
    """
    thresholds, warnings = from_mapping(read_existing(live_file).settings)
    assert warnings == []
    assert thresholds == DEFAULT_THRESHOLDS


def test_editing_a_floor_changes_how_stocks_are_categorised(live_file):
    """The whole point. He moves one number in Excel and the next refresh obeys it."""
    ws, start, _end = _band_rows(live_file, "score_bands")
    wb = ws.parent
    ws.cell(row=start, column=2, value=400)   # "Legendary" now starts at 400
    wb.save(live_file)

    thresholds, warnings = from_mapping(read_existing(live_file).settings)
    assert warnings == []
    assert thresholds.score_bands.label_for(393.0) == "Exceptional"
    assert thresholds.score_bands.label_for(401.0) == "Legendary"


def test_he_can_rename_a_category_without_breaking_the_read(live_file):
    """He renamed two between v2.0 and v2.1, so this is not hypothetical. The reader
    keys on the hidden column, never on the words he is free to change."""
    ws, start, _end = _band_rows(live_file, "value_bands")
    ws.cell(row=start, column=1, value="Once In A Lifetime")
    ws.parent.save(live_file)

    thresholds, warnings = from_mapping(read_existing(live_file).settings)
    assert thresholds.value_bands.label_for(17.4) == "Once In A Lifetime"
    assert warnings == []


def test_he_can_add_a_category_by_typing_into_a_spare_row(live_file):
    """Excel refuses to insert a row into a protected sheet, so the sheet ships spares.

    He types into whichever spare he lands on -- here the second, not the first -- and
    the band sorts to where its number puts it.
    """
    ws, _start, end = _band_rows(live_file, "value_bands")
    ws.cell(row=end - 2, column=1, value="Beyond Generational")
    ws.cell(row=end - 2, column=2, value=15.0)
    ws.parent.save(live_file)

    thresholds, warnings = from_mapping(read_existing(live_file).settings)
    assert warnings == []
    assert thresholds.value_bands.bands[0].label == "Beyond Generational"
    assert thresholds.value_bands.label_for(17.4) == "Beyond Generational"
    assert thresholds.value_bands.label_for(11.0) == "Generational"


def test_he_can_delete_a_category_by_clearing_its_two_cells(live_file):
    ws, start, _end = _band_rows(live_file, "score_bands")
    # Assigning .value, not passing value=None to cell(): openpyxl treats the keyword
    # form as "leave it alone", so the cell would never actually clear.
    ws.cell(row=start, column=1).value = None
    ws.cell(row=start, column=2).value = None
    ws.parent.save(live_file)

    thresholds, _ = from_mapping(read_existing(live_file).settings)
    assert "Legendary" not in [b.label for b in thresholds.score_bands.bands]
    assert thresholds.score_bands.label_for(393.0) == "Exceptional"


def test_a_typo_in_one_band_costs_that_band_and_not_the_refresh(live_file):
    ws, start, _end = _band_rows(live_file, "score_bands")
    ws.cell(row=start, column=2, value="one hundred and fifty")
    ws.parent.save(live_file)

    thresholds, warnings = from_mapping(read_existing(live_file).settings)
    assert len(warnings) == 1
    assert len(thresholds.score_bands.bands) == len(DEFAULT_SCORE_BANDS.bands) - 1


def test_a_scalar_setting_round_trips(live_file):
    _settings_cell(live_file, "mover_pe_pct", 15)
    thresholds, warnings = from_mapping(read_existing(live_file).settings)
    assert thresholds.mover_pe_pct == 15.0
    assert warnings == []


def test_first_run_with_no_existing_file_is_not_an_error(tmp_path):
    result = read_existing(tmp_path / "does-not-exist.xlsx")
    assert result.universe == []
    assert result.settings == {}


# --- Refusing to destroy data ---------------------------------------------------------


def test_unreadable_workbook_aborts_rather_than_overwriting(tmp_path):
    """A file that exists but cannot be parsed must never be silently replaced."""
    corrupt = tmp_path / "AJZ Dashboard.xlsx"
    corrupt.write_bytes(b"this is not a spreadsheet")
    with pytest.raises(WorkbookReadError):
        read_existing(corrupt)


def test_a_foreign_workbook_at_the_target_path_aborts(tmp_path):
    """Some other .xlsx sitting where ours goes must not be clobbered. This guard used
    to key off the Conviction sheet; Universe is the marker now."""
    foreign = tmp_path / "AJZ Dashboard.xlsx"
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.save(foreign)
    with pytest.raises(WorkbookReadError, match="Universe"):
        read_existing(foreign)


# --- Archiving the conviction scores we are deleting ----------------------------------


def _with_conviction_sheet(path):
    wb = load_workbook(path)
    ws = wb.create_sheet("Conviction")
    ws.cell(row=1, column=1, value="Ticker")
    ws.cell(row=2, column=1, value="NVDA")
    ws.cell(row=2, column=3, value=4)
    wb.save(path)


def test_a_legacy_conviction_sheet_is_archived_before_it_is_dropped(live_file, tmp_path):
    """Jeff asked for conviction to go. His scores were five human judgements per stock
    that no API can regenerate, and backups prune at thirty -- so they need somewhere
    permanent to land before the first post-upgrade refresh removes the sheet."""
    _with_conviction_sheet(live_file)
    archived = archive_conviction(live_file, tmp_path / "archive")

    assert archived is not None and archived.exists()
    saved = load_workbook(archived)
    assert saved.sheetnames == ["Conviction"]
    assert saved["Conviction"].cell(row=2, column=1).value == "NVDA"


def test_archiving_runs_once_and_never_overwrites_the_first_copy(live_file, tmp_path):
    """The second refresh must not replace the archive with a Conviction-free workbook."""
    _with_conviction_sheet(live_file)
    first = archive_conviction(live_file, tmp_path / "archive")
    assert archive_conviction(live_file, tmp_path / "archive") is None
    assert first.exists()


def test_a_workbook_with_no_conviction_sheet_archives_nothing(live_file, tmp_path):
    assert archive_conviction(live_file, tmp_path / "archive") is None


def test_archiving_never_raises_because_it_must_not_block_the_refresh(tmp_path):
    """This is insurance on a deletion he asked for. Failing to take a copy is bad;
    failing to refresh because the copy failed would be worse."""
    corrupt = tmp_path / "AJZ Dashboard.xlsx"
    corrupt.write_bytes(b"not a spreadsheet")
    assert archive_conviction(corrupt, tmp_path / "archive") is None
    assert archive_conviction(tmp_path / "missing.xlsx", tmp_path / "archive") is None


# --- Universe -------------------------------------------------------------------------


def test_universe_round_trips(live_file):
    universe = read_existing(live_file).universe
    assert {e.ticker for e in universe} >= {"NVDA", "TSM", "AVGO"}
    assert all(e.active for e in universe)


def test_setting_active_to_no_is_respected(live_file):
    wb = load_workbook(live_file)
    ws = wb["Universe"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "BE":
            ws.cell(row=row, column=4, value="NO")
    wb.save(live_file)

    entry = next(e for e in read_existing(live_file).universe if e.ticker == "BE")
    assert entry.active is False


def test_blank_active_cell_defaults_to_included(live_file):
    """A stock Jeff typed in without filling Active must appear, not vanish."""
    wb = load_workbook(live_file)
    ws = wb["Universe"]
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value="ASML")
    wb.save(live_file)

    entry = next(e for e in read_existing(live_file).universe if e.ticker == "ASML")
    assert entry.active is True


# --- Backups and atomic writes --------------------------------------------------------


def test_backup_is_written_before_replacement(live_file, tmp_path):
    backups = tmp_path / "backups"
    made = backup_workbook(live_file, backups, datetime(2026, 8, 19, 6, 5))
    assert made is not None and made.exists()
    assert len(read_existing(made).universe) > 0


def test_backups_are_pruned_to_the_keep_limit(live_file, tmp_path):
    backups = tmp_path / "backups"
    for hour in range(35):
        backup_workbook(live_file, backups, datetime(2026, 8, 19, 0, 0, hour))
    assert len(list(backups.glob("*.xlsx"))) == 30


def test_atomic_save_leaves_no_temp_file_behind(live_file):
    atomic_save(build_workbook(sample_stocks()), live_file)
    assert not live_file.with_suffix(".xlsx.tmp").exists()


def test_atomic_save_does_not_corrupt_the_target_on_failure(live_file, monkeypatch):
    """If the replace fails, the file Jeff opens must still be the old, valid one."""
    original = len(read_existing(live_file).universe)

    def boom(self, target):
        raise PermissionError("file is open in Excel")

    monkeypatch.setattr("pathlib.Path.replace", boom)
    with pytest.raises(WorkbookLockedError, match="open in Excel"):
        atomic_save(build_workbook(sample_stocks()), live_file)

    assert len(read_existing(live_file).universe) == original
    assert not live_file.with_suffix(".xlsx.tmp").exists()


# --- The colour round-trip ------------------------------------------------------------
#
# Jeff colour-coded the category names on the Settings sheet to make the tables visual,
# saved, and refreshed. The colours were gone -- the workbook is rebuilt from nothing on
# every refresh, so anything not explicitly read back is destroyed. He had found the same
# trap once before with his hand-entered tables. These tests are the fence around it.


def _fill_band_name(path, attr, offset, argb):
    """Simulate Jeff filling a category name cell with a colour."""
    ws, start, _end = _band_rows(path, attr)
    ws.cell(row=start + offset, column=1).fill = PatternFill("solid", fgColor=argb)
    ws.parent.save(path)


def test_a_colour_jeff_puts_on_a_category_survives_the_refresh(live_file):
    """The bug he reported, stated as a test."""
    _fill_band_name(live_file, "score_bands", 0, "FFB7791F")

    thresholds, warnings = from_mapping(read_existing(live_file).settings)

    assert warnings == []
    assert thresholds.score_bands.band_for("Legendary").color == "FFB7791F"


def test_an_untouched_sheet_reports_no_colours(live_file):
    """The seeded legend must read back as unset.

    The Settings sheet is seeded with each band's current colour so it doubles as the
    legend. If the reader could not tell that seed from a choice of Jeff's, our ramp
    would freeze on the first refresh and stop re-stretching when he adds a category.
    """
    thresholds, _warnings = from_mapping(read_existing(live_file).settings)
    assert all(band.color is None for band in thresholds.score_bands.bands)


def test_the_ramp_still_re_stretches_after_a_seeded_refresh(live_file):
    """Adding a category to a sheet that has already been seeded must re-shade the rest.

    This is the failure the seed could have caused: seven bands painted with a
    seven-band ramp, then an eighth added and eight categories sharing seven colours.
    """
    ws, _start, end = _band_rows(live_file, "score_bands")
    ws.cell(row=end - 1, column=1, value="Utterly Dead")   # a spare row
    ws.cell(row=end - 1, column=2, value=-50)
    ws.parent.save(live_file)

    thresholds, warnings = from_mapping(read_existing(live_file).settings)
    assert warnings == []
    assert len(thresholds.score_bands.bands) == 8

    build_workbook(sample_stocks(), thresholds=thresholds).save(live_file)
    ws, start, _end = _band_rows(live_file, "score_bands")
    seeded = [ws.cell(row=start + n, column=1).fill.fgColor.rgb for n in range(8)]
    assert len(set(seeded)) > 1, "eight categories collapsed onto one colour"


def test_a_colour_follows_a_category_he_renames(live_file):
    """Colour lives on the row, not on the word, because he re-words categories."""
    _fill_band_name(live_file, "value_bands", 0, "FF7B341E")
    ws, start, _end = _band_rows(live_file, "value_bands")
    ws.cell(row=start, column=1, value="Once In A Lifetime")
    ws.parent.save(live_file)

    thresholds, _warnings = from_mapping(read_existing(live_file).settings)
    assert thresholds.value_bands.band_for("Once In A Lifetime").color == "FF7B341E"


def test_his_colour_reaches_the_sheets_he_actually_reads(live_file):
    """Settings is where he sets it; Top Rankings and the Matrix are where he sees it.

    Colouring a cell on a settings page that changes nothing elsewhere would be a worse
    outcome than the bug -- it would look fixed and not be.
    """
    _fill_band_name(live_file, "value_bands", 0, "FF7B341E")
    thresholds, _warnings = from_mapping(read_existing(live_file).settings)
    build_workbook(sample_stocks(), thresholds=thresholds).save(live_file)

    wb = load_workbook(live_file)
    top = wb["Top Rankings"]
    labels = [top.cell(row=r, column=10) for r in range(2, top.max_row + 1)]
    matched = [c for c in labels if c.value == "Generational"]
    assert matched, "no stock is in the band under test"
    assert all(c.fill.fgColor.rgb == "FF7B341E" for c in matched)

    matrix = wb["Opportunity Matrix"]
    heads = [matrix.cell(row=5, column=c) for c in range(2, matrix.max_column + 1)]
    header = next(c for c in heads if c.value == "Generational")
    assert header.fill.fgColor.rgb == "FF7B341E"


def test_a_dark_colour_gets_readable_text(live_file):
    """He picks the fill; he should not also have to pick text that can be read on it."""
    _fill_band_name(live_file, "score_bands", 0, "FF0B1F3A")
    thresholds, _warnings = from_mapping(read_existing(live_file).settings)
    build_workbook(sample_stocks(), thresholds=thresholds).save(live_file)

    ws, start, _end = _band_rows(live_file, "score_bands")
    name = ws.cell(row=start, column=1)
    assert name.fill.fgColor.rgb == "FF0B1F3A"
    assert name.font.color.rgb == theme.INK_INVERSE
