"""Reading a colour back out of a cell Jeff filled in by hand.

Jeff coloured the category names on the Settings sheet to make them visual, saved, and
refreshed -- and the colours were gone, because every refresh rebuilds the workbook from
nothing. These tests cover the read-back that makes his fill survive, and they exist
mainly to pin down two things Excel does that are easy to get wrong:

* the top row of Excel's colour picker ("Theme Colors") does NOT store an RGB value. It
  stores a slot number into the workbook theme plus a tint. Resolving those is the
  difference between the feature working and appearing to ignore half the palette.
* a dark fill needs light text. Jeff picking navy must not produce black-on-navy.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill

from ajz import theme
from ajz.palette import apply_tint, ink_for, resolve_fill


def _roundtrip(fill: PatternFill | None):
    """Write a fill, save, reload the way store.py does, hand back (cell, workbook)."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Legendary"
    if fill is not None:
        ws["A1"].fill = fill
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    reloaded = load_workbook(buffer, data_only=True, read_only=True)
    return reloaded[reloaded.sheetnames[0]]["A1"], reloaded


def _assert_close(actual: str, expected: str, tolerance: int = 3) -> None:
    """Compare two hex colours channel by channel, allowing a rounding wobble.

    Excel converts through an integer HLS space and truncates; we go through floats. The
    two disagree by a unit or two on some tints. That is invisible on screen and it is
    stable -- the resolved colour is written back as plain RGB on the next refresh, so
    the difference cannot accumulate. Chasing bit-exactness with Excel's integer rounding
    would buy nothing anyone can see.
    """
    pairs = [(int(actual[i:i + 2], 16), int(expected[i:i + 2], 16)) for i in (0, 2, 4)]
    assert all(abs(a - e) <= tolerance for a, e in pairs), (
        f"{actual} is not within {tolerance}/255 per channel of {expected}")


class TestResolveFill:
    def test_reads_an_rgb_fill_through_a_read_only_load(self):
        """The load in store.py is read_only; styles have to survive it."""
        cell, wb = _roundtrip(PatternFill("solid", fgColor="FF2A78D6"))
        assert resolve_fill(cell, wb) == "FF2A78D6"

    def test_an_unfilled_cell_is_none(self):
        """No fill means "use our ramp", not "use white"."""
        cell, wb = _roundtrip(None)
        assert resolve_fill(cell, wb) is None

    def test_reads_a_six_digit_fill(self):
        cell, wb = _roundtrip(PatternFill("solid", fgColor="2A78D6"))
        assert resolve_fill(cell, wb) == "FF2A78D6"

    def test_resolves_a_theme_colour(self):
        """Excel's top picker row. Slot 4 is accent1.

        The literal below is accent1 of the Office theme openpyxl bundles. Jeff's file
        is saved by Excel and carries a newer theme whose accent1 is 4472C4, so the value
        is not the point -- the slot lookup is. What this pins down is that slot 4 finds
        accent1 in whatever theme the workbook actually carries.
        """
        cell, wb = _roundtrip(PatternFill("solid", fgColor=Color(theme=4, tint=0.0)))
        assert resolve_fill(cell, wb) == "FF4F81BD"

    def test_resolves_a_tinted_theme_colour(self):
        """"Accent 1, Lighter 40%" -- one click below the top row.

        The five rows under each theme colour in the picker are all the same slot with a
        different tint, so ignoring tint would collapse sixty swatches into twelve.
        """
        cell, wb = _roundtrip(PatternFill("solid", fgColor=Color(theme=4, tint=0.4)))
        resolved = resolve_fill(cell, wb)
        assert resolved is not None
        assert resolved != "FF4F81BD", "the tint was ignored"
        _assert_close(resolved[2:], "95B3D7")

    def test_theme_slots_zero_and_one_are_not_swapped(self):
        """Excel's index 0 is the light background, though the XML lists dk1 first."""
        cell, wb = _roundtrip(PatternFill("solid", fgColor=Color(theme=0, tint=0.0)))
        assert resolve_fill(cell, wb) == "FFFFFFFF"

    def test_an_unresolvable_colour_falls_back_to_none(self):
        """Never guess. A colour we cannot read means the ramp, not a wrong colour."""
        cell, wb = _roundtrip(PatternFill("solid", fgColor=Color(theme=99, tint=0.0)))
        assert resolve_fill(cell, wb) is None


class TestApplyTint:
    def test_zero_tint_is_the_identity(self):
        assert apply_tint("4472C4", 0.0) == "4472C4"

    def test_positive_tint_lightens(self):
        """The reference value is what Excel itself renders for "Lighter 40%"."""
        _assert_close(apply_tint("4472C4", 0.4), "8EAADB")

    def test_negative_tint_darkens(self):
        _assert_close(apply_tint("4472C4", -0.5), "223963")

    def test_white_cannot_lighten_further(self):
        assert apply_tint("FFFFFF", 0.5) == "FFFFFF"


class TestInkFor:
    @pytest.mark.parametrize("dark", ["FF0D366B", "FF2A78D6", "FF000000"])
    def test_dark_fills_take_light_ink(self, dark):
        assert ink_for(dark) == theme.INK_INVERSE

    @pytest.mark.parametrize("light", ["FFCDE2FB", "FFFFFF00", "FFFFFFFF"])
    def test_light_fills_take_dark_ink(self, light):
        assert ink_for(light) == theme.INK_PRIMARY

    def test_no_fill_takes_the_default_ink(self):
        assert ink_for(None) == theme.INK_PRIMARY
