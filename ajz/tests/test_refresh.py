"""End-to-end refresh tests — the real code path Jeff's machine runs daily.

These replace the hand-rolled reconstruction in test_store.py: the conviction round-trip
is exercised through `refresh()` itself, so the test proves the shipping path preserves
his scores rather than proving a test helper does.
"""

from __future__ import annotations

from datetime import datetime, timedelta

import pytest
from openpyxl import load_workbook

from ajz.bands import Band
from ajz.calc import rank_stocks
from ajz.fixtures import sample_stocks
from ajz.settings import DEFAULT_THRESHOLDS, from_mapping
from ajz.refresh import (
    AuthError,
    FetchResult,
    QuotaError,
    fetcher_from_fixtures,
    refresh,
)
from ajz.status import RefreshState
from ajz.store import WorkbookReadError, read_existing

MONDAY = datetime(2026, 8, 17, 6, 0)
NEXT_MONDAY = MONDAY + timedelta(days=7)


@pytest.fixture
def paths(tmp_path):
    return {
        "workbook_path": tmp_path / "AJZ Dashboard.xlsx",
        "history_path": tmp_path / "history.sqlite",
        "backup_dir": tmp_path / "backups",
    }


@pytest.fixture
def seed():
    return [
        type("E", (), {"ticker": s.data.ticker, "company": s.data.company,
                       "sector": s.data.sector, "active": True, "notes": None})()
        for s in sample_stocks()
    ]


def run(paths, when=MONDAY, fetch=None, **kw):
    from ajz.store import UniverseEntry

    universe = [
        UniverseEntry(ticker=s.data.ticker, company=s.data.company, sector=s.data.sector)
        for s in sample_stocks()
    ]
    return refresh(
        fetch=fetch or fetcher_from_fixtures(),
        seed_universe=universe,
        now=when,
        **paths,
        **kw,
    )


def _set_band(path, attr, index, label=None, floor=None):
    """Simulate Jeff editing one row of a category table on the Settings sheet.

    Note the explicit `.value =` assignment. `ws.cell(row, col, value=None)` silently
    SKIPS the assignment in openpyxl, so passing None through the constructor would
    leave the old value in place and quietly make a clearing test pass.
    """
    wb = load_workbook(path)
    ws = wb["Settings"]
    first = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value == f"table:{attr}":
            first = row + 2
            break
    assert first is not None, f"table:{attr} not found"
    ws.cell(row=first + index, column=1).value = label
    ws.cell(row=first + index, column=2).value = floor
    wb.save(path)


# --- First run ------------------------------------------------------------------------


def test_first_run_creates_a_usable_workbook(paths):
    outcome = run(paths)
    assert outcome.written
    assert paths["workbook_path"].exists()
    assert outcome.status.state is RefreshState.OK
    assert len(outcome.ranked) > 0


def test_first_run_fires_no_movement_alerts(paths):
    """With no history, nothing has 'moved'. A wall of first-run alerts trains Jeff
    to ignore alerts entirely."""
    outcome = run(paths)
    movement = [a.value for s in outcome.stocks for a in s.alerts
                if a.value in {"UPGRADE", "DOWNGRADE"}]
    assert movement == []


# --- Jeff's edits surviving the real path ---------------------------------------------


def test_jeffs_category_tables_survive_a_refresh(paths):
    """THE critical guarantee, now that his tables are the hand-entered data.

    Before v2.1 this test protected conviction scores. It protects the same thing: work
    Jeff typed, which the refresh rewrites the file over the top of. What is irreplaceable
    changed; that it must never be silently lost did not.
    """
    run(paths)
    _set_band(paths["workbook_path"], "value_bands", 0, "Once In A Lifetime", 12.0)

    run(paths, when=NEXT_MONDAY)

    thresholds, _ = from_mapping(read_existing(paths["workbook_path"]).settings)
    assert thresholds.value_bands.bands[0].label == "Once In A Lifetime"
    assert thresholds.value_bands.bands[0].floor == 12.0


def test_nothing_is_lost_across_many_refreshes(paths):
    """Ten refreshes must not erode his edits. Slow leaks are the dangerous kind."""
    run(paths)
    _set_band(paths["workbook_path"], "score_bands", 0, "Untouchable", 200.0)

    for week in range(1, 11):
        run(paths, when=MONDAY + timedelta(days=7 * week))

    thresholds, _ = from_mapping(read_existing(paths["workbook_path"]).settings)
    assert thresholds.score_bands.bands[0] == Band("Untouchable", 200.0)


def test_an_edited_table_actually_changes_the_categories(paths):
    """Surviving is not enough — the refresh has to obey it."""
    outcome = run(paths)
    assert _label_of(outcome, "NVDA") == "Generational"

    _set_band(paths["workbook_path"], "value_bands", 0, "Generational", 20.0)
    outcome = run(paths, when=NEXT_MONDAY)
    assert _label_of(outcome, "NVDA") == "Elite"


def test_a_legacy_conviction_sheet_is_archived_on_the_upgrade_refresh(paths, tmp_path):
    """Upgrading from v1.x deletes a sheet full of his hand-entered judgements. It gets
    copied somewhere permanent first, and he is told where."""
    run(paths)
    wb = load_workbook(paths["workbook_path"])
    ws = wb.create_sheet("Conviction")
    ws.cell(row=1, column=1, value="Ticker")
    ws.cell(row=2, column=1, value="NVDA")
    ws.cell(row=2, column=3, value=4)
    wb.save(paths["workbook_path"])

    outcome = run(paths, when=NEXT_MONDAY)

    assert outcome.written
    assert any("conviction" in w.lower() for w in outcome.warnings)
    assert list(tmp_path.glob("*conviction scores (archived).xlsx"))


def test_a_backup_is_written_on_every_refresh_after_the_first(paths):
    run(paths)
    outcome = run(paths, when=NEXT_MONDAY)
    assert outcome.backup is not None and outcome.backup.exists()


# --- Failure handling -----------------------------------------------------------------


def test_auth_failure_leaves_the_existing_workbook_untouched(paths):
    run(paths)
    _set_band(paths["workbook_path"], "value_bands", 0, "Mine", 12.0)
    before = paths["workbook_path"].read_bytes()

    def failing(tickers):
        raise AuthError("401")

    outcome = run(paths, when=NEXT_MONDAY, fetch=failing)

    assert outcome.written is False
    assert outcome.status.state is RefreshState.AUTH_ERROR
    assert paths["workbook_path"].read_bytes() == before  # byte-identical


def test_quota_failure_does_not_blank_the_workbook(paths):
    run(paths)

    def over_quota(tickers):
        raise QuotaError("429")

    outcome = run(paths, when=NEXT_MONDAY, fetch=over_quota)
    assert outcome.written is False
    assert len(read_existing(paths["workbook_path"]).universe) > 0


def test_partial_fetch_is_labelled_and_lists_the_missing(paths):
    def partial(tickers):
        full = fetcher_from_fixtures()(tickers)
        return FetchResult(stocks=full.stocks[:-2], missing=("XYZ", "ABC"))

    outcome = run(paths, fetch=partial)
    assert outcome.status.state is RefreshState.PARTIAL
    assert "XYZ" in outcome.status.note


def test_corrupt_existing_workbook_aborts_without_writing(paths):
    paths["workbook_path"].write_bytes(b"not a spreadsheet")
    with pytest.raises(WorkbookReadError):
        run(paths)
    assert paths["workbook_path"].read_bytes() == b"not a spreadsheet"


def test_empty_universe_is_refused_rather_than_producing_a_blank_file(paths):
    from ajz.refresh import refresh as do_refresh

    with pytest.raises(ValueError, match="No active tickers"):
        do_refresh(fetch=fetcher_from_fixtures(), seed_universe=[], now=MONDAY, **paths)


# --- History integration --------------------------------------------------------------


def test_a_big_score_move_produces_an_alert_on_the_second_run(paths):
    """Jeff's rule: more than 25% on the AJZ Score. The alert v5.1 could never fire,
    because the data did not exist anywhere in the workbook."""
    import sqlite3

    run(paths)
    outcome = run(paths, when=NEXT_MONDAY, snapshot=False)
    subject = rank_stocks(outcome.stocks)[0].ticker

    # Halve last week's stored score, so this week reads as a large climb.
    with sqlite3.connect(paths["history_path"]) as conn:
        conn.execute(
            "UPDATE snapshots SET ajz_score = ajz_score / 2, "
            "ajz_value_score = ajz_value_score / 2 "
            "WHERE ticker = ? AND snapshot_date = ?",
            (subject, MONDAY.date().isoformat()),
        )

    outcome = run(paths, when=NEXT_MONDAY)
    moved = next(s for s in outcome.stocks if s.ticker == subject)
    assert any(a.value == "UPGRADE" for a in moved.alerts)


def test_the_movers_sheet_reports_that_move(paths):
    """The sheet Jeff said "doesn't look like its updating" — because it was a stub that
    never wrote a row no matter how many refreshes ran."""
    import sqlite3
    from ajz.history import History
    from ajz.refresh import movement_report
    from ajz.settings import DEFAULT_THRESHOLDS

    run(paths)
    outcome = run(paths, when=NEXT_MONDAY, snapshot=False)
    subject = rank_stocks(outcome.stocks)[0].ticker

    with sqlite3.connect(paths["history_path"]) as conn:
        conn.execute(
            "UPDATE snapshots SET ajz_score = ajz_score / 2, "
            "ajz_value_score = ajz_value_score / 2 "
            "WHERE ticker = ? AND snapshot_date = ?",
            (subject, MONDAY.date().isoformat()),
        )

    report = movement_report(outcome.stocks, History(paths["history_path"]),
                             NEXT_MONDAY.date(), DEFAULT_THRESHOLDS)
    assert report["has_baseline"] is True
    assert any(r["ticker"] == subject and r["what"] == "AJZ Score"
               for r in report["rows"])


def test_the_first_refresh_says_no_baseline_rather_than_no_movers(paths):
    """"Nothing moved" and "I have nothing to compare against" are different claims,
    and only one of them is true on a first run."""
    from ajz.history import History
    from ajz.refresh import movement_report
    from ajz.settings import DEFAULT_THRESHOLDS

    outcome = run(paths)
    report = movement_report(outcome.stocks, History(paths["history_path"]),
                             MONDAY.date(), DEFAULT_THRESHOLDS)
    assert report["has_baseline"] is False
    assert report["rows"] == []


def test_a_quiet_week_reports_no_movers_with_a_baseline(paths):
    from ajz.history import History
    from ajz.refresh import movement_report
    from ajz.settings import DEFAULT_THRESHOLDS

    run(paths)
    outcome = run(paths, when=NEXT_MONDAY)
    report = movement_report(outcome.stocks, History(paths["history_path"]),
                             NEXT_MONDAY.date(), DEFAULT_THRESHOLDS)
    assert report["has_baseline"] is True
    assert report["rows"] == []


def test_history_survives_workbook_regeneration(paths):
    from ajz.history import History

    run(paths)
    run(paths, when=NEXT_MONDAY)
    assert len(History(paths["history_path"]).snapshot_dates()) == 2


# --- Seeding the first run ------------------------------------------------------------


def test_seeded_first_run_is_immediately_useful(paths):
    """REGRESSION (v5.1): Jeff's first open must not be an empty grid.

    It no longer can be. Conviction was the one input he had to supply by hand before
    anything was classified, and removing it means every stock the API returns is fully
    categorised on the very first refresh, with nothing to fill in.
    """
    outcome = run(paths)
    ranked = rank_stocks(outcome.stocks)
    assert len(ranked) > 10
    assert all(s.value_label for s in ranked)
    assert all(s.score_label for s in ranked)


# --- Settings round-trip (spec §6.5) --------------------------------------------------


def _set_setting(path, key, value):
    """Simulate Jeff editing the Settings sheet, matching on the hidden key column."""
    wb = load_workbook(path)
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value == key:
            ws.cell(row=row, column=2).value = value
            break
    else:
        raise AssertionError(f"setting {key!r} not found")
    wb.save(path)


def _label_of(outcome, ticker):
    return next(s for s in outcome.stocks if s.ticker == ticker).value_label


def test_a_threshold_edit_survives_the_refresh(paths):
    """His edit must still be there after the rewrite. That it also *takes effect* is
    proved by test_an_edited_table_actually_changes_the_categories."""
    run(paths)
    _set_setting(paths["workbook_path"], "mover_pe_pct", 15.0)

    run(paths, when=NEXT_MONDAY)

    assert read_existing(paths["workbook_path"]).settings["mover_pe_pct"] == 15.0


def test_a_bad_setting_falls_back_without_stopping_the_refresh(paths):
    """A typo must cost a default, not a morning's dashboard."""
    run(paths)
    _set_setting(paths["workbook_path"], "warning_value", "five")

    outcome = run(paths, when=NEXT_MONDAY)
    assert outcome.written is True
    assert any("not a number" in w for w in outcome.warnings)


def test_a_bad_band_falls_back_without_stopping_the_refresh(paths):
    run(paths)
    _set_band(paths["workbook_path"], "score_bands", 0, "Legendary", "one fifty")

    outcome = run(paths, when=NEXT_MONDAY)
    assert outcome.written is True
    assert any("not a number" in w for w in outcome.warnings)


def test_clearing_a_setting_restores_its_default(paths):
    run(paths)
    _set_setting(paths["workbook_path"], "mover_pe_pct", 15.0)
    run(paths, when=NEXT_MONDAY)

    _set_setting(paths["workbook_path"], "mover_pe_pct", None)
    run(paths, when=NEXT_MONDAY + timedelta(days=7))

    thresholds, _ = from_mapping(read_existing(paths["workbook_path"]).settings)
    assert thresholds.mover_pe_pct == 10.0


def test_clearing_a_whole_table_restores_the_shipped_one(paths):
    """He must be able to get back to Jeff's own numbers without reinstalling."""
    run(paths)
    _set_band(paths["workbook_path"], "value_bands", 0, "Nonsense", 99.0)
    run(paths, when=NEXT_MONDAY)

    for index in range(8):
        _set_band(paths["workbook_path"], "value_bands", index, None, None)
    run(paths, when=NEXT_MONDAY + timedelta(days=7))

    thresholds, _ = from_mapping(read_existing(paths["workbook_path"]).settings)
    assert thresholds.value_bands == DEFAULT_THRESHOLDS.value_bands


def test_settings_and_universe_edits_coexist(paths):
    """Both editable sheets must survive the same rewrite."""
    run(paths)
    _set_setting(paths["workbook_path"], "buy_value", 4.0)
    _set_band(paths["workbook_path"], "score_bands", 0, "Untouchable", 200.0)

    run(paths, when=NEXT_MONDAY)

    saved = read_existing(paths["workbook_path"])
    thresholds, _ = from_mapping(saved.settings)
    assert thresholds.buy_value == 4.0
    assert thresholds.score_bands.bands[0] == Band("Untouchable", 200.0)
    assert len(saved.universe) > 10


def test_a_workbook_without_a_settings_sheet_still_refreshes(paths):
    """Older workbooks predate this sheet; defaults are the right answer for them."""
    run(paths)
    wb = load_workbook(paths["workbook_path"])
    del wb["Settings"]
    wb.save(paths["workbook_path"])

    outcome = run(paths, when=NEXT_MONDAY)
    assert outcome.written is True


def test_a_falling_category_does_not_fire_an_upgrade(paths):
    """REGRESSION (found on live data): a stock dropping a category fired both alerts."""
    import sqlite3

    run(paths)
    outcome = run(paths, when=NEXT_MONDAY, snapshot=False)
    subject = rank_stocks(outcome.stocks)[0].ticker

    # Make last week's snapshot claim it sat in the very best band, so this week is a fall.
    with sqlite3.connect(paths["history_path"]) as conn:
        conn.execute(
            "UPDATE snapshots SET category = 'Generational' "
            "WHERE ticker = ? AND snapshot_date = ?",
            (subject, MONDAY.date().isoformat()),
        )
        conn.execute(
            "UPDATE snapshots SET ajz_score = ? WHERE ticker = ? AND snapshot_date = ?",
            (next(s.ajz_score for s in outcome.stocks if s.ticker == subject),
             subject, MONDAY.date().isoformat()),
        )

    outcome = run(paths, when=NEXT_MONDAY)
    moved = next(s for s in outcome.stocks if s.ticker == subject)
    values = {a.value for a in moved.alerts}
    assert not ("UPGRADE" in values and "DOWNGRADE" in values), \
        f"{subject} contradicts itself: {sorted(values)}"
