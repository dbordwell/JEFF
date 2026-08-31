"""Tests for the workbook generator.

These assert the properties that make the file safe in a non-technical user's hands:
no formulas, correct sheets protected, validation present, and no phantom empty rows.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from ajz.fixtures import sample_stocks
from ajz.status import RefreshState, RefreshStatus
from ajz.workbook import build_workbook

EXPECTED_SHEETS = [
    "Dashboard", "Top Rankings", "Opportunity Matrix",
    "Alerts", "Movers", "Universe", "Settings",
]
EDITABLE_SHEETS = {"Universe", "Settings"}


@pytest.fixture(scope="module")
def stocks():
    return sample_stocks()


@pytest.fixture(scope="module")
def saved(stocks):
    """Round-trip through bytes, so we test the real file rather than the object graph."""
    buffer = BytesIO()
    build_workbook(stocks).save(buffer)
    buffer.seek(0)
    return load_workbook(buffer)


def test_has_exactly_the_expected_sheets(saved):
    assert saved.sheetnames == EXPECTED_SHEETS


def test_no_reported_value_comes_from_a_formula(saved):
    """REGRESSION (v5.1): the workbook was formula-driven and could go #REF!.

    Every number Jeff reads is computed in Python and written as a static value, so
    nothing recalculates behind him and nothing breaks when he sorts or edits.

    The one exception, added in v2.1, is the "Range (automatic)" column on the Settings
    sheet — and it is an exception to the letter of this rule rather than its point. That
    column displays the band boundaries implied by the floors he types, live, as he types
    them. Nothing reads it back: classification uses the floors. If the formula broke he
    would lose a hint, not a number. v5.1's formulas computed the numbers themselves,
    which is what made #REF! a data-loss event rather than a cosmetic one.
    """
    offenders = []
    for ws in saved.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                if ws.title == "Settings" and cell.column == 3:
                    continue
                offenders.append(f"{ws.title}!{cell.coordinate}={cell.value}")
    assert offenders == [], f"formulas found: {offenders}"


def test_the_settings_range_column_is_locked_so_he_cannot_break_it(saved):
    """He is meant to edit the floors beside it, and those cells are adjacent."""
    ws = saved["Settings"]
    formulas = [c for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert formulas, "the Range column wrote no formulas at all"
    assert all(c.protection.locked is not False for c in formulas)


def test_display_sheets_are_protected(saved):
    for name in EXPECTED_SHEETS:
        if name not in EDITABLE_SHEETS:
            assert saved[name].protection.sheet is True, f"{name} is not protected"


def _band_row(ws, attr, offset=0):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=4).value == f"table:{attr}":
            return row + 2 + offset
    raise AssertionError(f"table:{attr} not found")


def test_band_name_and_floor_cells_are_unlocked(saved):
    """Protection is on for the sheet, but his category tables must be typeable."""
    ws = saved["Settings"]
    row = _band_row(ws, "score_bands")
    assert ws.cell(row=row, column=1).protection.locked is False
    assert ws.cell(row=row, column=2).protection.locked is False


def test_every_table_ships_blank_spare_rows_so_he_can_add_a_category(saved):
    """Excel refuses to insert a row into a protected sheet. Without spares, "add a
    band" is impossible in the file rather than merely awkward — and he demonstrably
    adds bands: v2.0 -> v2.1 added two."""
    from ajz.settings import BAND_TABLES, SPARE_BAND_ROWS

    ws = saved["Settings"]
    for attr, _title, default in BAND_TABLES:
        first = _band_row(ws, attr)
        spares = 0
        for offset in range(len(default.bands), len(default.bands) + SPARE_BAND_ROWS):
            cell = ws.cell(row=first + offset, column=1)
            assert cell.value is None
            assert cell.protection.locked is False
            spares += 1
        assert spares == SPARE_BAND_ROWS


def test_settings_tables_are_delimited_so_spare_rows_stay_spares(saved):
    """A blank row must not be mistaken for the end of a table, or a band typed into the
    second spare would be silently dropped."""
    from ajz.settings import BAND_TABLES

    ws = saved["Settings"]
    markers = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert markers.count("table:end") == len(BAND_TABLES)


def test_universe_active_column_is_a_yes_no_list(saved):
    ws = saved["Universe"]
    lists = [v for v in ws.data_validations.dataValidation if v.type == "list"]
    assert lists and "YES,NO" in lists[0].formula1


def test_no_phantom_empty_rows_on_rankings(saved, stocks):
    """REGRESSION (v5.1): 499 pre-filled rows whose formulas returned 0.

    The rankings sheet must hold exactly one row per real stock, plus the header.
    """
    ws = saved["Top Rankings"]
    assert ws.max_row == len(stocks) + 1


def test_unrankable_stocks_are_shown_but_not_ranked(saved):
    """RIVN is loss-making: it must appear, with no rank number."""
    ws = saved["Top Rankings"]
    rows = {ws.cell(row=r, column=2).value: r for r in range(2, ws.max_row + 1)}
    assert "RIVN" in rows
    assert ws.cell(row=rows["RIVN"], column=1).value == "—"


def test_ranked_stocks_are_in_descending_value_order(saved):
    ws = saved["Top Rankings"]
    values = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "—":
            break
        values.append(ws.cell(row=r, column=9).value)   # AJZ Value moved to column I
    assert values == sorted(values, reverse=True)


def test_missing_values_render_as_dash_never_zero(saved):
    """REGRESSION (v5.1): missing data displayed as 0, which reads as a real bad score."""
    ws = saved["Top Rankings"]
    rows = {ws.cell(row=r, column=2).value: r for r in range(2, ws.max_row + 1)}
    snow = rows["SNOW"]  # missing ROIC -> no AJZ Score
    assert ws.cell(row=snow, column=5).value == "—"
    assert ws.cell(row=snow, column=6).value == "—"


def test_notes_explain_why_a_stock_could_not_be_scored(saved):
    ws = saved["Top Rankings"]
    rows = {ws.cell(row=r, column=2).value: r for r in range(2, ws.max_row + 1)}
    note = ws.cell(row=rows["SNOW"], column=12).value
    assert "roic" in note.lower()


def test_an_expensive_stock_is_categorised_not_dropped(saved):
    """NET has sound fundamentals and a punishing P/E. It must land in the bottom band
    with a word for it, not vanish and not read as a data failure.

    This test used to assert NET showed "Needs Conviction" — a state that existed only
    because we required a hand-entered score before anything could be classified. With
    conviction gone there is no such limbo: everything the API returns is categorised.
    """
    ws = saved["Top Rankings"]
    rows = {ws.cell(row=r, column=2).value: r for r in range(2, ws.max_row + 1)}
    assert ws.cell(row=rows["NET"], column=10).value == "Expensive"
    assert ws.cell(row=rows["NET"], column=8).value == "Bubble"


# --- Status banner ------------------------------------------------------------------


def _banner_of(stocks, status):
    buffer = BytesIO()
    build_workbook(stocks, status=status).save(buffer)
    buffer.seek(0)
    return load_workbook(buffer)["Dashboard"].cell(row=4, column=2).value


def test_healthy_banner_states_the_date(stocks):
    when = datetime(2026, 8, 19, 6, 5)
    text = _banner_of(stocks, RefreshStatus(RefreshState.OK, data_as_of=when))
    assert "Data current as of" in text
    assert "2026" in text


def test_stale_banner_is_plain_english_and_dates_the_data(stocks):
    """Stale-but-labelled beats blank or wrong (spec §10)."""
    when = datetime(2026, 8, 18, 6, 5)
    text = _banner_of(stocks, RefreshStatus(RefreshState.STALE, data_as_of=when))
    assert "Could not reach the data provider" in text
    assert "18 Aug 2026" in text


def test_auth_error_tells_jeff_to_call_a_human(stocks):
    text = _banner_of(stocks, RefreshStatus(RefreshState.AUTH_ERROR))
    assert "call Dave" in text


def test_no_error_codes_or_jargon_in_any_banner(stocks):
    """Jeff must never see a status code, a URL, or the word 'API'."""
    jargon = ["API", "HTTP", "40", "50", "null", "None", "exception", "traceback"]
    for state in RefreshState:
        text = _banner_of(stocks, RefreshStatus(state, data_as_of=datetime(2026, 8, 19)))
        for term in jargon:
            assert term not in text, f"{state.value} banner leaked {term!r}: {text}"


def test_partial_refresh_lists_the_missing_tickers(stocks):
    status = RefreshStatus(
        RefreshState.PARTIAL, data_as_of=datetime(2026, 8, 19),
        missing_tickers=("XYZ", "ABC"),
    )
    buffer = BytesIO()
    build_workbook(stocks, status=status).save(buffer)
    buffer.seek(0)
    note = load_workbook(buffer)["Dashboard"].cell(row=5, column=2).value
    assert "2 tickers had no data today" in note
    assert "XYZ" in note and "ABC" in note


# --- Dashboard KPIs -----------------------------------------------------------------


def test_the_dashboard_is_emptied_but_still_carries_the_status_banner(saved):
    """Jeff, v2.1: "I would eliminate the data but leave the sheet for future use."

    The banner stays. It is not data in the sense he meant — it is the whole
    error-reporting surface, the one place that says whether these numbers arrived today
    or are three days stale. Removing it would leave a silent failure nowhere to appear.
    """
    ws = saved["Dashboard"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "Portfolio Quality Index" not in text
    assert "Average AJZ Value" not in text
    assert "reserved for a future" in text


def test_empty_universe_still_produces_a_valid_workbook(stocks):
    """It must degrade to an honest empty dashboard, never crash."""
    buffer = BytesIO()
    build_workbook([]).save(buffer)
    buffer.seek(0)
    assert load_workbook(buffer).sheetnames == EXPECTED_SHEETS


def test_every_ranked_row_carries_all_three_category_words(saved):
    """Jeff's v2.1 layout: each number is followed by the word for it, so he reads across
    a row instead of back and forth to a legend."""
    ws = saved["Top Rankings"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 13)]
    assert headers[4:10] == ["AJZ Score", "Score Category", "Forward P/E",
                             "P/E Category", "AJZ Value", "Value Category"]

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "—":
            continue   # unrankable rows legitimately have no Value category
        for col in (6, 8, 10):
            assert ws.cell(row=row, column=col).value not in (None, ""), \
                f"row {row} column {col} has no category word"


def test_no_conviction_survives_anywhere_in_the_file(saved):
    """Jeff: "Get Rid Of Conviction Calculation and references to same throughout."

    "Throughout" included a sheet, two columns, four Opportunity Matrix buckets and half
    of every alert rule. A leftover mention is a promise the rest of the file breaks.
    """
    assert "Conviction" not in saved.sheetnames
    for ws in saved.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert "onviction" not in cell.value, \
                        f"{ws.title}!{cell.coordinate}: {cell.value!r}"


# --- Sheet protection must not block the one edit we invited -------------------------


def test_settings_sheet_allows_jeff_to_change_a_fill():
    """v2.1 asked him to colour the category cells. Protection was refusing to let him.

    This is the gap that 304 passing tests could not see. openpyxl ignores sheet
    protection entirely -- it will happily write a fill into a cell that Excel would
    refuse -- so the colour round-trip tests proved the read-back worked while the
    gesture they depend on was blocked on his machine.

    In OOXML a sheetProtection attribute of "1" means the feature is PROTECTED, not
    permitted. formatCells defaulted to 1, so every fill command on the Settings sheet
    was refused. He reported it twice as "the colour didn't stick".
    """
    wb = build_workbook(sample_stocks())
    assert wb["Settings"].protection.formatCells is False


def test_the_generated_sheets_stay_locked_down():
    """Only Settings opens up, and only for formatting.

    Everything else is regenerated from scratch each refresh, so an edit there is work
    he would lose without being told. Settings is the one sheet whose formatting we
    read back, which is what makes it the one sheet worth unlocking.
    """
    wb = build_workbook(sample_stocks())
    for name in ("Top Rankings", "Opportunity Matrix", "Alerts", "Movers"):
        assert wb[name].protection.formatCells is not False, f"{name} was unlocked too"


def test_the_saved_file_really_tells_excel_formatting_is_allowed(tmp_path):
    """Assert on the bytes Excel reads, not on the openpyxl object.

    The object model is our side of the contract; the sheetProtection element is the
    side Excel enforces. They are not the same thing, and the whole defect was that the
    library let us write a fill the application would have refused.
    """
    import re
    import zipfile

    path = tmp_path / "AJZ Dashboard.xlsx"
    build_workbook(sample_stocks()).save(path)

    with zipfile.ZipFile(path) as archive:
        protections = [
            re.search(r"<sheetProtection[^>]*/>", archive.read(name).decode())
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet")
        ]

    found = [m.group(0) for m in protections if m]
    assert found, "no sheet is protected at all"
    permissive = [p for p in found if 'formatCells="0"' in p]
    assert len(permissive) == 1, (
        f"expected exactly one sheet to permit formatting, found {len(permissive)}"
    )
    assert all('sheet="1"' in p for p in found), "a sheet lost its protection entirely"
