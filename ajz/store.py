"""Reading Jeff's hand-entered data back out of the workbook (spec §7.3).

What he types is the only thing in this system that cannot be regenerated. The AJZ inputs
all come from an API; his universe and his category tables are judgements no endpoint can
supply. Losing them means he redoes work he may not remember doing.

So this module is written defensively and the invariant is absolute:

    NEVER write a workbook containing less of Jeff's own input than the one it replaces.

Every failure path here aborts rather than degrades. A refresh that does not happen is a
minor annoyance; a refresh that silently blanks his edits is unrecoverable.

Conviction was removed at his instruction in "Requested Changes for Items 2.1". Because
that deletes data he hand-entered, `archive_conviction` copies any existing Conviction
sheet somewhere permanent *before* the first post-upgrade refresh drops it. Rolling
backups would not do: they prune, so the scores would survive thirty refreshes and then
vanish. He said the calculation doesn't do anything; he did not say to burn his notes.
"""

from __future__ import annotations

import os
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from . import theme
from .palette import resolve_fill
from .settings import TABLE_END, TABLE_PREFIX

CONVICTION_SHEET = "Conviction"  # legacy; read once to archive, never written
UNIVERSE_SHEET = "Universe"
SETTINGS_SHEET = "Settings"
SETTINGS_KEY_COL = 4   # hidden column holding the field key
SETTINGS_VALUE_COL = 2

TICKER_COL = 1
UNIVERSE_ACTIVE_COL = 4
UNIVERSE_NOTES_COL = 5

BACKUPS_TO_KEEP = 30


class WorkbookReadError(RuntimeError):
    """Raised when an existing workbook cannot be read.

    Always fatal to a refresh. If we cannot prove what Jeff had, we must not overwrite it.
    """


@dataclass(frozen=True)
class UniverseEntry:
    ticker: str
    company: str | None = None
    sector: str | None = None
    active: bool = True
    notes: str | None = None


@dataclass(frozen=True)
class ReadResult:
    universe: list[UniverseEntry]
    warnings: tuple[str, ...] = ()
    settings: dict[str, object] = field(default_factory=dict)


def read_existing(path: Path) -> ReadResult:
    """Read Jeff's own edits -- universe and settings -- out of an existing workbook.

    On first run the file does not exist yet, which is a legitimate empty result rather
    than an error. Any *other* failure raises, because a workbook that exists but cannot
    be read is exactly the case where overwriting would destroy data.
    """
    if not path.exists():
        return ReadResult(universe=[])

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - any failure here must abort the refresh
        raise WorkbookReadError(
            f"Could not open the existing workbook at {path}: {exc}. "
            "Refusing to overwrite it."
        ) from exc

    warnings: list[str] = []
    try:
        universe = _read_universe_sheet(wb, warnings)
        settings = _read_settings_sheet(wb, warnings)
    except WorkbookReadError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise WorkbookReadError(
            f"Could not read saved data from {path}: {exc}. Refusing to overwrite it."
        ) from exc
    finally:
        wb.close()

    return ReadResult(universe=universe, warnings=tuple(warnings), settings=settings)


def archive_conviction(path: Path, archive_dir: Path) -> Path | None:
    """Preserve a legacy Conviction sheet before a refresh drops it. Idempotent.

    Returns the archive path if one was written, or None if there was nothing to save
    (no workbook, no Conviction sheet, or it has already been archived).

    Deliberately never raises: this is insurance on a deletion Jeff explicitly asked for,
    so failing to take a copy must not be able to block the refresh he did ask for.
    """
    if not path.exists():
        return None

    target = archive_dir / f"{path.stem} - conviction scores (archived).xlsx"
    if target.exists():
        return None

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001 - insurance must not become a failure mode
        return None
    try:
        if CONVICTION_SHEET not in wb.sheetnames:
            return None
    finally:
        wb.close()

    try:
        archive_dir.mkdir(parents=True, exist_ok=True)
        source = load_workbook(path)
        for name in list(source.sheetnames):
            if name != CONVICTION_SHEET:
                source.remove(source[name])
        source.save(target)
    except Exception:  # noqa: BLE001
        return None
    return target


def _read_universe_sheet(wb, warnings: list[str]) -> list[UniverseEntry]:
    if UNIVERSE_SHEET not in wb.sheetnames:
        # This used to key off the Conviction sheet. Universe is now the marker that says
        # "this file came from us" -- an existing file without one is somebody else's
        # spreadsheet at the same path, and overwriting it is not ours to do.
        raise WorkbookReadError(
            f"The workbook has no '{UNIVERSE_SHEET}' sheet. This is not a workbook this "
            "tool produced; refusing to overwrite it."
        )

    ws = wb[UNIVERSE_SHEET]
    seen: set[str] = set()
    out: list[UniverseEntry] = []

    for row in ws.iter_rows(min_row=2):
        if not row:
            continue
        raw_ticker = row[0].value if row else None
        if raw_ticker is None or not str(raw_ticker).strip():
            continue
        ticker = str(raw_ticker).strip().upper()
        if ticker in seen:
            warnings.append(f"{ticker}: duplicate row in Universe sheet; kept the first")
            continue
        seen.add(ticker)

        def cell(index: int):
            return row[index].value if len(row) > index else None

        active_raw = cell(UNIVERSE_ACTIVE_COL - 1)
        # Default to active: a blank Active cell should include the stock, not silently
        # drop it. Only an explicit "NO"/"FALSE"/0 removes one.
        active = str(active_raw).strip().upper() not in {"NO", "FALSE", "0", "N"} if active_raw is not None else True

        out.append(
            UniverseEntry(
                ticker=ticker,
                company=(str(cell(1)).strip() if cell(1) else None),
                sector=(str(cell(2)).strip() if cell(2) else None),
                active=active,
                notes=(str(cell(UNIVERSE_NOTES_COL - 1)).strip() if cell(UNIVERSE_NOTES_COL - 1) else None),
            )
        )

    return out


# --- Backups --------------------------------------------------------------------------


def backup_workbook(path: Path, backup_dir: Path, now: datetime) -> Path | None:
    """Copy the current workbook aside before it is replaced.

    Cheap insurance on the only irreplaceable data in the system. Runs on every refresh,
    not just risky ones, because the risky ones are the ones we did not predict.
    """
    if not path.exists():
        return None

    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = now.strftime("%Y%m%d-%H%M%S")
    target = backup_dir / f"{path.stem}-{stamp}.xlsx"
    shutil.copy2(path, target)
    _prune_backups(backup_dir)
    return target


def _prune_backups(backup_dir: Path, keep: int = BACKUPS_TO_KEEP) -> None:
    backups = sorted(backup_dir.glob("*.xlsx"), key=lambda p: p.name, reverse=True)
    for stale in backups[keep:]:
        stale.unlink(missing_ok=True)


# --- Safe write -----------------------------------------------------------------------


class WorkbookLockedError(RuntimeError):
    """The workbook is open in Excel. Retry later; never write around it."""


def atomic_save(wb, path: Path) -> None:
    """Write to a temp file in the same directory, then swap it in.

    Two failure modes this protects against:

    * A crash mid-write leaving a truncated workbook. The swap is atomic, so the file
      Jeff opens is always either the old one or the new one, never a partial one.
    * Excel holding the file open on Windows, which makes the replace fail with
      PermissionError. We abort and keep the good file rather than writing around it.

    The temp name carries the process id. A fixed name would be shared by two refreshes
    running at once, and they would interleave their writes into a single temp file that
    then gets atomically swapped into place — an atomic swap of a corrupt file. The
    lock in `lock.py` should stop that happening; this makes it harmless if it ever does.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(f"{path.suffix}.{os.getpid()}.tmp")

    try:
        wb.save(temp)
    except PermissionError as exc:
        temp.unlink(missing_ok=True)
        raise WorkbookLockedError(
            f"Could not write to {temp}. Is the dashboard open in Excel?"
        ) from exc

    try:
        temp.replace(path)
    except PermissionError as exc:
        temp.unlink(missing_ok=True)
        raise WorkbookLockedError(
            f"Could not replace {path}. The dashboard is probably open in Excel; "
            "the previous version has been left untouched."
        ) from exc


def _cell(row, index: int):
    """One cell of a row, or None past its end. Short rows are normal in openpyxl."""
    return row[index].value if len(row) > index else None


def _band_colour(wb, row) -> str | None:
    """The colour Jeff filled a band's name cell with, or None for "use our ramp".

    Our own seed colour reads back as None on purpose -- see `theme.is_ramp_colour`.
    Anything we cannot resolve to a real colour is also None: a band that keeps its old
    shading is a non-event, whereas a band painted a colour he did not choose is us
    making a claim about his own categories that he never made.
    """
    if not row:
        return None
    colour = resolve_fill(row[0], wb)
    return None if theme.is_ramp_colour(colour) else colour


def _read_settings_sheet(wb, warnings: list[str]) -> dict[str, object]:
    """Read Jeff's settings edits: scalar values and his three category tables.

    Matches on the hidden key column rather than the visible label, so re-wording a
    label in a future version cannot silently orphan a setting he has changed.

    The sheet holds two shapes. Scalars are `label | value | explanation` rows keyed by
    column D. Tables are announced by a `table:<field>` marker in column D, run as
    `band name | starts at` rows, and close on a `table:end` marker.

    Blank rows inside a table are spares, not terminators. The sheet is protected, so
    Excel will not let Jeff insert a row; blank spares are how he adds a band, and he
    will not reliably fill the topmost one. That is why the table needs an explicit end
    marker rather than ending at the first gap.

    A missing sheet is NOT an error: workbooks written before Settings existed are still
    perfectly valid, and defaults are the right answer for them.
    """
    if SETTINGS_SHEET not in wb.sheetnames:
        return {}

    ws = wb[SETTINGS_SHEET]
    out: dict[str, object] = {}
    table_rows: list[tuple[object, object, str | None]] | None = None
    skip_header = False

    for row in ws.iter_rows(min_row=2):
        marker = _cell(row, SETTINGS_KEY_COL - 1)
        marker = str(marker).strip() if marker is not None else ""

        if marker == TABLE_END:
            table_rows = None
            continue

        if marker.startswith(TABLE_PREFIX):
            table_rows = []
            out[marker] = table_rows
            skip_header = True   # the row after the marker is the column header
            continue

        if table_rows is not None:
            if skip_header:
                skip_header = False
                continue
            label, floor = _cell(row, 0), _cell(row, 1)
            if (label is None or not str(label).strip()) and floor is None:
                continue  # an unused spare row
            table_rows.append((label, floor, _band_colour(wb, row)))
            continue

        if not marker:
            continue
        value = _cell(row, SETTINGS_VALUE_COL - 1)
        if value is None or (isinstance(value, str) and not value.strip()):
            continue  # cleared cell -> fall back to the default for that field
        out[marker] = value

    return out
